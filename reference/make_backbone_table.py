"""Create reference/lt_backbone_comparison.xlsx from paper_summary_en.xlsx data."""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── color palette ──────────────────────────────────────────────────────────────
NAVY_FILL  = PatternFill(fill_type="solid", fgColor="002F4F8F")
NAVY_FONT  = Font(bold=True, size=10, color="FFFFFF")
HDR_ALIGN  = Alignment(wrap_text=True, horizontal="center", vertical="center")

EVEN_FILL  = PatternFill(fill_type="solid", fgColor="00EEF2FF")
ODD_FILL   = PatternFill(fill_type=None)
DATA_FONT  = Font(bold=False, size=9)
DATA_ALIGN = Alignment(wrap_text=True, horizontal="center", vertical="top")

THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

GRAY_FILL  = PatternFill(fill_type="solid", fgColor="F2F2F2")

# backbone category fills
CAT_FILLS = {
    "CNN-ResNet":  PatternFill(fill_type="solid", fgColor="D6E4F0"),
    "CNN-ResNeXt": PatternFill(fill_type="solid", fgColor="BDD7EE"),
    "ViT":         PatternFill(fill_type="solid", fgColor="E2EFDA"),
    "SNN":         PatternFill(fill_type="solid", fgColor="FFE699"),
    "NAS":         PatternFill(fill_type="solid", fgColor="FCE4D6"),
    "Mixed":       PatternFill(fill_type="solid", fgColor="EAD1DC"),
}

# ── paper data ─────────────────────────────────────────────────────────────────
# Keys: abb, title, src, mtype, c10, c100, inet, inat, bcat, note
# "-" means not evaluated on that benchmark
PAPERS = [
    dict(
        abb="IMMAX",
        title="Balancing Scales: Theoretical Framework for Imbalanced Classification",
        src="ICML 2025",
        mtype="Loss function",
        c10="ResNet-32",
        c100="ResNet-32",
        inet="-",
        inat="-",
        bcat="CNN-ResNet",
        note="CIFAR not specified as LT; standard imbalanced split (IR=10/50/100)",
    ),
    dict(
        abb="ImOOD",
        title="Towards Balanced OOD Detection under Long-tailed Training Data",
        src="NeurIPS 2024",
        mtype="OOD detection",
        c10="ResNet-32",
        c100="ResNet-32",
        inet="ResNeXt-50",
        inat="-",
        bcat="Mixed",
        note="OOD sets: SVHN, LSUN, iSUN; OOD bias correction on LT training",
    ),
    dict(
        abb="PRL",
        title="Controllable Long-tail Learning via Preference Representation Learning",
        src="NeurIPS 2024",
        mtype="Expert ensemble (hypernetwork)",
        c10="ResNet-32",
        c100="ResNet-32",
        inet="ResNeXt-50",
        inat="-",
        bcat="Mixed",
        note="Dirichlet preference vectors; flexible head-tail trade-off at inference",
    ),
    dict(
        abb="NCMC",
        title="Neural Collapse To Multiple Centers For Imbalanced Data",
        src="NeurIPS 2024",
        mtype="Loss function (NC-based)",
        c10="ResNet-32",
        c100="ResNet-32",
        inet="-",
        inat="-",
        bcat="CNN-ResNet",
        note="Balanced (not LT) CIFAR; fixed ETF classifier variant",
    ),
    dict(
        abb="LLM-AutoDA",
        title="LLM-Driven Automatic Data Augmentation for Long-tailed Problems",
        src="NeurIPS 2024",
        mtype="Augmentation (LLM-guided)",
        c10="ResNet-32",
        c100="ResNet-32",
        inet="ResNeXt-50",
        inat="-",
        bcat="Mixed",
        note="LLM proposes class-wise DA strategies; iterative refinement",
    ),
    dict(
        abb="LOS",
        title="Rethinking Classifier Re-Training: Label Over-Smooth Can Balance",
        src="ICLR 2025",
        mtype="Decoupled training (Stage-2 loss)",
        c10="-",
        c100="ResNet-32",
        inet="ResNeXt-50",
        inat="ResNet-50",
        bcat="Mixed",
        note="Plug-in Stage-2 smoother; no class-frequency priors needed",
    ),
    dict(
        abb="ConMix",
        title="Contrastive Mixup at Representation Level for Long-tailed Deep Clustering",
        src="ICLR 2025",
        mtype="Contrastive clustering",
        c10="ResNet-32",
        c100="ResNet-32",
        inet="ResNeXt-50",
        inat="-",
        bcat="Mixed",
        note="Deep clustering; multiple IR settings tested",
    ),
    dict(
        abb="LFN",
        title="Learning from Neighbors: Category Extrapolation for Long-Tail Learning",
        src="CVPR 2025",
        mtype="Augmentation (web crawl)",
        c10="-",
        c100="-",
        inet="ResNeXt-50",
        inat="ResNet-50",
        bcat="Mixed",
        note="Auxiliary open-set categories from LLM + web crawl; neighbor-silencing loss",
    ),
    dict(
        abb="TAET",
        title="Two-Stage Adversarial Equalization Training on Long-Tailed Distributions",
        src="CVPR 2025",
        mtype="Adversarial training",
        c10="ResNet-32",
        c100="ResNet-32",
        inet="-",
        inat="-",
        bcat="CNN-ResNet",
        note="Balanced adversarial robustness; CE stabilization + HARL",
    ),
    dict(
        abb="MGS",
        title="Long-Tailed Classification with Multi-Granularity Semantics",
        src="ICCV 2025",
        mtype="Representation (LLM-guided)",
        c10="ResNet-32",
        c100="ResNet-32",
        inet="ResNeXt-50",
        inat="-",
        bcat="Mixed",
        note="Semantic knowledge graph from LLM descriptions; SKCL loss",
    ),
    dict(
        abb="TinyChange",
        title="LT Class-Incremental Learning via Geometric Prototype Alignment",
        src="ICCV 2025",
        mtype="Continual / CIL",
        c10="-",
        c100="ResNet-32",
        inet="ResNet-50",
        inat="-",
        bcat="CNN-ResNet",
        note="LT-CIL focus; geometric init on unit hypersphere",
    ),
    dict(
        abb="DBM",
        title="Difficulty-aware Balancing Margin Loss for Long-tailed Recognition",
        src="AAAI 2025",
        mtype="Loss function",
        c10="ResNet-32",
        c100="ResNet-32",
        inet="ResNeXt-50",
        inat="ResNet-50",
        bcat="Mixed",
        note="Class-wise + instance-wise dual margin; plug-in loss",
    ),
    dict(
        abb="BCE3S",
        title="BCE-Based Tripartite Synergistic Learning for Long-tailed Recognition",
        src="AAAI 2026",
        mtype="Loss function (BCE-based)",
        c10="ResNet-32",
        c100="ResNet-32",
        inet="ResNeXt-50",
        inat="ResNet-50",
        bcat="Mixed",
        note="Three BCE objectives jointly; decouples classifier vectors",
    ),
    dict(
        abb="IP2SL",
        title="Informative and Positive-Balanced Sampling for Long-tail Learning",
        src="NeurIPS 2025",
        mtype="Decoupled training (sampler)",
        c10="ResNet-32",
        c100="ResNet-32",
        inet="ResNeXt-50",
        inat="-",
        bcat="Mixed",
        note="BNS + IP-DPP sampling; MI-maximizing tail oversampling",
    ),
    dict(
        abb="IP-DPP",
        title="Long-Tailed Recognition via Information-Preservable Two-Stage Learning",
        src="NeurIPS 2025",
        mtype="Decoupled training (Stage-2)",
        c10="ResNet-32",
        c100="ResNet-32",
        inet="ResNeXt-50",
        inat="ResNet-50",
        bcat="Mixed",
        note="DPP-based info-preserving balanced sampling in Stage 2",
    ),
    dict(
        abb="IBCL",
        title="Improved Balanced Classification with Theoretically Grounded Loss Functions",
        src="NeurIPS 2025",
        mtype="Loss function (theory)",
        c10="ResNet-32",
        c100="ResNet-32",
        inet="ResNeXt-50",
        inat="-",
        bcat="Mixed",
        note="GLA + GCA losses; H-consistency bound derivation",
    ),
    dict(
        abb="LTRL",
        title="Boosting Long-Tail Recognition via Reflective Learning",
        src="ECCV 2024",
        mtype="Training strategy (plug-in)",
        c10="ResNet-32",
        c100="ResNet-32",
        inet="ResNeXt-50",
        inat="ResNet-50",
        bcat="Mixed",
        note="Reflection buffer + cross-class feature + gradient harmonizer",
    ),
    dict(
        abb="DaSC",
        title="Distribution-Aware Robust Learning from Long-Tailed Data with Noisy Labels",
        src="ECCV 2024",
        mtype="Noisy label + LT",
        c10="ResNet-32",
        c100="ResNet-32",
        inet="ResNet-18 (mini)",
        inat="-",
        bcat="CNN-ResNet",
        note="Simultaneous LT + label noise; DaCC centroid + SBCL loss",
    ),
    dict(
        abb="LT-SpikeFormer",
        title="Tackling Long-Tailed Data Challenges in SNNs via Heterogeneous KD",
        src="IJCAI 2025",
        mtype="Architecture (SNN + KD)",
        c10="SpikingResNet",
        c100="SpikingResNet",
        inet="SpikeFormer",
        inat="-",
        bcat="SNN",
        note="SNN teacher-student KD; only SNN LT paper; IR=10/50/100",
    ),
    dict(
        abb="RLAL",
        title="Revisiting Long-Tailed Learning: Insights from an Architectural Perspective",
        src="CIKM 2025",
        mtype="Architecture (NAS)",
        c10="LT-DARTS",
        c100="LT-DARTS",
        inet="LT-DARTS",
        inat="-",
        bcat="NAS",
        note="Architecture-level LT study; LT-AggConv + LT-HierConv search ops",
    ),
]

# ── column definitions ──────────────────────────────────────────────────────────
COLS = [
    ("Abb.", 8),
    ("Title", 42),
    ("Source", 12),
    ("Method Type", 22),
    ("CIFAR-10-LT\nBackbone", 14),
    ("CIFAR-100-LT\nBackbone", 14),
    ("ImageNet-LT\nBackbone", 14),
    ("iNat-2018\nBackbone", 14),
    ("Backbone\nCategory", 14),
    ("Notes", 38),
]

def apply_cell(cell, font, fill, alignment, border):
    cell.font = font
    if fill is not None:
        cell.fill = fill
    cell.alignment = alignment
    cell.border = border


def make_sheet1(wb):
    ws = wb.create_sheet("LT Benchmark Comparison")

    # header
    for col_i, (hdr, width) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=col_i, value=hdr)
        apply_cell(cell, NAVY_FONT, NAVY_FILL, HDR_ALIGN, BORDER)
        ws.column_dimensions[get_column_letter(col_i)].width = width
    ws.row_dimensions[1].height = 36

    # data rows
    for row_i, p in enumerate(PAPERS, start=2):
        row_fill = EVEN_FILL if (row_i % 2 == 0) else ODD_FILL
        vals = [
            p["abb"], p["title"], p["src"], p["mtype"],
            p["c10"], p["c100"], p["inet"], p["inat"],
            p["bcat"], p["note"],
        ]
        for col_i, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            col_hdr = COLS[col_i - 1][0]
            # backbone category cells get category color
            if "Backbone\nCategory" in col_hdr:
                fill = CAT_FILLS.get(p["bcat"], row_fill)
            else:
                fill = row_fill
            apply_cell(cell, DATA_FONT, fill, DATA_ALIGN, BORDER)
        ws.row_dimensions[row_i].height = 60

    ws.freeze_panes = "A2"
    return ws


def make_legend(wb):
    ws = wb.create_sheet("범례 (Legend)")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40

    # header
    for col_i, hdr in enumerate(["Backbone Category", "Description"], start=1):
        cell = ws.cell(row=1, column=col_i, value=hdr)
        apply_cell(cell, NAVY_FONT, NAVY_FILL, HDR_ALIGN, BORDER)
    ws.row_dimensions[1].height = 28

    legends = [
        ("CNN-ResNet",  "#D6E4F0 (sky blue) — ResNet-32, ResNet-50, ResNet-18, ResNet-152"),
        ("CNN-ResNeXt", "#BDD7EE (steel blue) — ResNeXt-50"),
        ("ViT",         "#E2EFDA (light green) — Vision Transformer variants"),
        ("SNN",         "#FFE699 (yellow) — Spiking Neural Networks (SpikingResNet, SpikeFormer)"),
        ("NAS",         "#FCE4D6 (peach) — Neural Architecture Search (LT-DARTS)"),
        ("Mixed",       "#EAD1DC (rose) — Multiple backbones across datasets"),
    ]
    for row_i, (cat, desc) in enumerate(legends, start=2):
        cat_cell = ws.cell(row=row_i, column=1, value=cat)
        cat_fill = CAT_FILLS.get(cat, EVEN_FILL)
        apply_cell(cat_cell, DATA_FONT, cat_fill, DATA_ALIGN, BORDER)

        desc_cell = ws.cell(row=row_i, column=2, value=desc)
        apply_cell(desc_cell, DATA_FONT, EVEN_FILL if row_i % 2 == 0 else ODD_FILL, DATA_ALIGN, BORDER)
        ws.row_dimensions[row_i].height = 28

    return ws


def make_summary(wb):
    ws = wb.create_sheet("데이터셋별 논문 수")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 40

    for col_i, hdr in enumerate(["Dataset", "# Papers", "Abbreviations"], start=1):
        cell = ws.cell(row=1, column=col_i, value=hdr)
        apply_cell(cell, NAVY_FONT, NAVY_FILL, HDR_ALIGN, BORDER)
    ws.row_dimensions[1].height = 28

    datasets = [
        ("CIFAR-10-LT",    "c10"),
        ("CIFAR-100-LT",   "c100"),
        ("ImageNet-LT",    "inet"),
        ("iNaturalist-2018", "inat"),
    ]

    for row_i, (name, key) in enumerate(datasets, start=2):
        papers_with = [p for p in PAPERS if p[key] != "-"]
        count = len(papers_with)
        abbs  = ", ".join(p["abb"] for p in papers_with)

        row_fill = EVEN_FILL if row_i % 2 == 0 else ODD_FILL
        for col_i, val in enumerate([name, count, abbs], start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            apply_cell(cell, DATA_FONT, row_fill, DATA_ALIGN, BORDER)
        ws.row_dimensions[row_i].height = 48

    return ws


# ── build workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)          # remove default sheet

make_sheet1(wb)
make_legend(wb)
make_summary(wb)

out_path = "reference/lt_backbone_comparison.xlsx"
wb.save(out_path)
print(f"Saved {out_path} ({len(PAPERS)} papers, 3 sheets)")
